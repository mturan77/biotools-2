import streamlit as st
import pandas as pd
import time
import datetime
import io
import re
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import ChromeType
from Bio import SeqIO

# Excel Stil Kütüphaneleri
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Page Config ---
st.set_page_config(page_title="SMART Pro Analyzer", layout="wide", initial_sidebar_state="expanded")

# --- CSS ---
st.markdown("""
<style>
    .stApp {background-color: #f8f9fa;}
    code {font-family: 'Consolas', monospace !important; font-size: 0.8rem;}
    div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] {
        background-color: white;
        border-radius: 5px;
    }
</style>
""", unsafe_allow_html=True)

# --- Session State (Hafıza) ---
if 'analysis_complete' not in st.session_state:
    st.session_state.analysis_complete = False
if 'final_excel_data' not in st.session_state:
    st.session_state.final_excel_data = None
if 'final_df' not in st.session_state:
    st.session_state.final_df = None
if 'filename' not in st.session_state:
    st.session_state.filename = "SMART_Results.xlsx"
if 'logs' not in st.session_state:
    st.session_state.logs = []
if 'queue_df' not in st.session_state:
    st.session_state.queue_df = None

# --- Driver ---
def get_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless") 
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    try:
        service = Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install())
        return webdriver.Chrome(service=service, options=chrome_options)
    except Exception as e:
        st.error(f"Driver Error: {e}")
        return None

# --- Helpers ---
def extract_number(text):
    match = re.search(r'\d+', text)
    if match: return int(match.group())
    return None

def update_queue_display(placeholder, df):
    placeholder.dataframe(df, use_container_width=True, hide_index=True)

def reset_analysis():
    """Analizi sıfırlar (Callback içinden çağrılmaz, akıştan çağrılır)"""
    st.session_state.analysis_complete = False
    st.session_state.final_excel_data = None
    st.session_state.final_df = None
    st.session_state.logs = []
    st.session_state.queue_df = None

# --- SIDEBAR TASARIMI ---
with st.sidebar:
    st.header("1. Input Configuration")
    
    # Dosya yükleme
    uploaded_file = st.file_uploader("Upload Source File (FASTA)", type=["fa", "fasta", "txt"], key="file_uploader")
    
    st.divider()
    
    # Bilgilendirme
    if not st.session_state.analysis_complete:
        st.info("**ℹ️ Ready:** Upload FASTA and click Start.")
        start_btn = st.button("🚀 Start Analysis", type="primary", use_container_width=True)
        new_btn = False
    else:
        st.success("**✅ Complete:** Download results below.")
        start_btn = False # İşlem bittiği için Start butonu gizlenir
        # New Analysis butonu BURADA çıkar
        if st.button("🔄 New Analysis", type="secondary", use_container_width=True):
            reset_analysis()
            st.rerun()

# --- ANA EKRAN TASARIMI ---
st.title("🧬 SMART Database: Genomic Data Acquisition Protocol")
st.markdown("Automated High-Throughput Retrieval System with **Multi-Color Grouping**.")
st.divider()

# --- Main Logic ---

def log(msg, level="INFO", placeholder=None):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    icon = "ℹ️" if level=="INFO" else "✅" if level=="SUCCESS" else "⚠️" if level=="WARNING" else "❌"
    st.session_state.logs.insert(0, f"[{ts}] {icon} {msg}")
    if placeholder:
        placeholder.code("\n".join(st.session_state.logs), language="bash")

# 1. DURUM: Analiz Başlatıldı
if start_btn and uploaded_file and not st.session_state.analysis_complete:
    
    stringio = io.StringIO(uploaded_file.getvalue().decode("utf-8"))
    sequences = list(SeqIO.parse(stringio, "fasta"))
    
    # Prefix
    file_prefix = "SMART"
    if len(sequences) > 0:
        first_id = sequences[0].id
        file_prefix = first_id[:4] if len(first_id) >= 4 else first_id
    
    col_queue, col_log = st.columns([1.5, 1])
    
    # Queue
    queue_data = [{"Accession ID": s.id, "Status": "QUEUED", "Domains": 0} for s in sequences]
    df_queue = pd.DataFrame(queue_data)
    st.session_state.queue_df = df_queue
    
    with col_queue:
        st.subheader("📋 Processing Queue")
        q_place = st.empty()
        update_queue_display(q_place, df_queue)
        
    with col_log:
        st.subheader(f"📟 System Telemetry ({file_prefix})")
        log_cont = st.container(height=450)
        with log_cont:
            log_place = st.empty()

    log(f"Pipeline initialized. Target sequences: {len(sequences)}", "INFO", log_place)
    
    driver = get_driver()
    all_results = []
    
    if driver:
        log("Headless WebDriver connection established.", "SUCCESS", log_place)
        
        for idx, seq_record in enumerate(sequences):
            prot_id = seq_record.id
            prot_seq = str(seq_record.seq)
            
            df_queue.loc[df_queue['Accession ID'] == prot_id, 'Status'] = '⏳ PROCESSING'
            st.session_state.queue_df = df_queue
            update_queue_display(q_place, df_queue)
            
            log(f"Processing Accession: {prot_id} ({idx+1}/{len(sequences)})", "INFO", log_place)
            
            try:
                driver.get("https://smart.embl-heidelberg.de/")
                try:
                    driver.implicitly_wait(1)
                    mode = driver.find_elements(By.CSS_SELECTOR, "a[href*='change_mode.cgi?mode=normal']")
                    if mode: driver.execute_script("arguments[0].click();", mode[0])
                except: pass
                
                try:
                    pfam = driver.find_element(By.XPATH, "//input[contains(@name, 'PFAM') or @name='DO_PFAM']")
                    if not pfam.is_selected(): driver.execute_script("arguments[0].click();", pfam)
                except: pass
                
                seq_in = driver.find_element(By.NAME, "SEQUENCE")
                seq_in.clear()
                seq_in.send_keys(prot_seq)
                driver.execute_script("arguments[0].form.submit();", seq_in)
                
                start_t = time.time()
                found = False
                valid = 0
                
                while time.time() - start_t < 60:
                    try:
                        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable tbody tr")))
                        rows = driver.find_elements(By.CSS_SELECTOR, "table.dataTable tbody tr")
                        
                        if len(rows) > 0 and "No data available" not in rows[0].text:
                            for row in rows:
                                cols = row.find_elements(By.TAG_NAME, "td")
                                if len(cols) >= 5: continue 
                                if len(cols) >= 3:
                                    name = cols[0].text.strip()
                                    if not name:
                                        try: name = cols[0].find_element(By.TAG_NAME, "a").text.strip()
                                        except: pass
                                    if name.lower() in ["coiled coil", "low complexity"]: continue
                                    s_val = extract_number(cols[1].text)
                                    e_val = extract_number(cols[2].text)
                                    eval_txt = cols[3].text.strip() if len(cols)>3 else "N/A"
                                    if s_val is not None:
                                        all_results.append({
                                            "Protein_ID": prot_id,
                                            "Feature": name,
                                            "Start": s_val,
                                            "End": e_val,
                                            "E-value": eval_txt
                                        })
                                        valid += 1
                            found = True
                            break
                        if "No domains found" in driver.page_source:
                            found = True
                            break
                    except: time.sleep(1)
                
                if found:
                    df_queue.loc[df_queue['Accession ID'] == prot_id, 'Status'] = '✅ COMPLETED'
                    df_queue.loc[df_queue['Accession ID'] == prot_id, 'Domains'] = valid
                    if valid > 0: log(f"Extraction successful. {valid} valid domains recorded.", "SUCCESS", log_place)
                    else: log("No confident domains found.", "WARNING", log_place)
                else:
                    df_queue.loc[df_queue['Accession ID'] == prot_id, 'Status'] = '❌ TIMEOUT'
                    
            except Exception as e:
                df_queue.loc[df_queue['Accession ID'] == prot_id, 'Status'] = '❌ FAILED'
                log(f"Runtime Error: {str(e)[:40]}", "ERROR", log_place)
            
            st.session_state.queue_df = df_queue
            update_queue_display(q_place, df_queue)

        driver.quit()
        log(f"Analysis Protocol Completed for {file_prefix}.", "SUCCESS", log_place)
        
        # EXCEL OLUŞTURMA
        if all_results:
            df_final = pd.DataFrame(all_results)
            df_final = df_final.sort_values(by=["Protein_ID", "Start"])
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='SMART_Domains')
                ws = writer.sheets['SMART_Domains']
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                hex_colors = ["E2EFDA", "DDEBF7", "FCE4D6", "FFF2CC", "E4DFEC", "EDEDED", "D0E0E3"]
                fills = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in hex_colors]
                
                current_protein = None
                color_index = 0
                
                for row in ws.iter_rows(min_row=2, max_col=5):
                    protein_cell = row[0]
                    if protein_cell.value != current_protein:
                        current_protein = protein_cell.value
                        color_index = (color_index + 1) % len(fills)
                    current_fill = fills[color_index]
                    for cell in row:
                        cell.fill = current_fill
                        cell.border = Border(left=Side(style='thin', color="BFBFBF"), 
                                             right=Side(style='thin', color="BFBFBF"),
                                             bottom=Side(style='thin', color="BFBFBF"))
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 4

            st.session_state.final_excel_data = output.getvalue()
            st.session_state.final_df = df_final
            st.session_state.filename = f"{file_prefix}_SMART_Results.xlsx"
            st.session_state.analysis_complete = True
            st.rerun()

# 2. DURUM: Analiz Bitti (Sonuçları Göster)
if st.session_state.analysis_complete and st.session_state.final_df is not None:
    
    col_queue, col_log = st.columns([1.5, 1])
    with col_queue:
        st.subheader("📋 Processing Queue (Completed)")
        if st.session_state.queue_df is not None:
             update_queue_display(st.empty(), st.session_state.queue_df)
    with col_log:
        st.subheader("📟 System Telemetry")
        log_cont = st.container(height=450)
        with log_cont:
             st.code("\n".join(st.session_state.logs), language="bash")

    st.divider()
    st.success("✅ Analysis Completed & Cached.")
    st.subheader("📊 Final Organized Dataset")
    st.dataframe(st.session_state.final_df, use_container_width=True)
    
    st.download_button(
        label=f"📥 Download Report ({st.session_state.filename})",
        data=st.session_state.final_excel_data,
        file_name=st.session_state.filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )

elif not start_btn and not st.session_state.analysis_complete:
    st.info("👈 Please upload a protein FASTA file and click 'Start' to begin.")
